from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
##work on already created
# wb=load_workbook('Book1.xlsx')
# ws= wb['Sheet1']

# print(ws['A2'].value)
# wb.create_sheet("Test")
# print(wb.sheetnames)
# #re-write
# ws['A3'].value="New"
##row entry
# ws.append(['avit', 'is', 'great'])
# wb.save('Book1.xlsx')

# #create new wookbook
# wb= Workbook()

wb=load_workbook('Book1.xlsx')
#give defualt sheet
ws=wb.active

#accessing row and colum
# for row in range(1,11):
#     for col in range(1,5):
#         char = get_column_letter(col)
#         ws[char +str(row)]= char + str(row)
# #        print(ws[char+str(row)].value)
##title of sheet
#ws.title="Data"

#merge cells
ws.merge_cells("A1:D2")
#ws.unmerge_cells("A1:D1")


## insert or delete 
# ws.insert_rows(5)
# ws.insert_cols(3)
# ws.delete_rows(7)

##move  data row or column
#ws.move_range("C1:D11", rows=2, cols=2)
wb.save('Book1.xlsx')